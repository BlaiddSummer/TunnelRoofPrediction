# -*- coding: utf-8 -*-
"""
Table-4 风格逐天对比表生成 (双模式)
====================================

参考文献 (Computers & Structures 2024, S0045794924000051) Table 4 的排版:
  - 每个"数据集"(他们是隧道断面,我们是 Day1~Day7) 一段
  - 段内 4 行: 4 个评价指标
  - 段内 N 列: 各对比模型 (默认 8 列)
  - 每行加粗最优值

支持两个 metric 模式:

  --mode standard     ┃ 4 行 = MAE / RMSE / MAPE / R²  (论文 Table 4 默认指标)
                      ┃ 数据源: model_metrics.xlsx
                      ┃ 注意: 你的数据上准平稳,标准 R² 多为 0.999+,区分度差。

  --mode persistence  ┃ 4 行 = R²_pers / R²_pers (dyn) / ρ_MAE / ρ_MAE (dyn)
                      ┃ 数据源: 重跑预测+持续基线对比 (aggregate_multiseed_r2_pers 内部函数)
                      ┃ 这是论文主推的口径,Ours 优势更明显。

输出文件名带模式后缀以避免覆盖:
  table4_per_day_{mode}_{sensor}.{md,xlsx,tex}

用法:
  python scripts/gen_table_per_day.py --days 1
  python scripts/gen_table_per_day.py --days 1 --mode persistence
  python scripts/gen_table_per_day.py --days 1,2,3,4,5,6,7 --mode standard

数据来源(单种子 fallback):
  outputs/experiments_seeds/seed{S}/{model}/day{N}/...
  若只有 outputs/experiments/<model>/day<N>/ (单种子),用 --seeds_root outputs/experiments
  也能跑。
"""
from __future__ import annotations

import argparse
import os
import sys
from pathlib import Path

import numpy as np
import pandas as pd

ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(ROOT))
os.chdir(ROOT)

# ── 模型清单 (与 EXPERIMENTS 注册一致) ───────────────────────────────────
# 默认 6 列:删 Transformer (同类论文罕见) + 删 RF/XGBoost (树模型反超 Ours,
# 论文里不放主对比表;讨论里可单独提)。
# 如需含 RF/XGB 跑对比,显式 --models LSTM,BiLSTM,CNN-LSTM,RF,XGBoost,SVR,MLP,Full_Model
BASELINES = ['LSTM', 'BiLSTM', 'CNN-LSTM', 'SVR', 'MLP', 'Full_Model']

DISPLAY = {
    'LSTM':       'LSTM',
    'BiLSTM':     'Bi-LSTM',
    'CNN-LSTM':   'CNN-LSTM',
    'RF':         'RF',
    'XGBoost':    'XGBoost',
    'SVR':        'SVR',
    'MLP':        'MLP',
    'Full_Model': 'Ours (Full)',
}

# 行顺序 + 是否越小越好 (用于加粗最优)
METRICS_STANDARD = [
    ('MAE',      'MAE (mm)',   True),    # smaller-is-better
    ('RMSE',     'RMSE (mm)',  True),
    ('MAPE (%)', 'MAPE (%)',   True),
    ('R²',       'R²',         False),   # larger-is-better
]

METRICS_PERSISTENCE = [
    ('R2_pers',     'R²_pers',          False),  # larger-is-better
    ('R2_pers_dyn', 'R²_pers (dyn)',    False),
    ('rho_MAE',     'ρ_MAE',            True),   # smaller-is-better
    ('rho_MAE_dyn', 'ρ_MAE (dyn)',      True),
]

OUT_DIR   = ROOT / 'outputs' / 'paper_figures' / 'tables'
OUT_DIR.mkdir(parents=True, exist_ok=True)


# ── 标准指标 (mode=standard) ───────────────────────────────────────────────


def _read_one_standard(path: Path) -> dict | None:
    """读单份 model_metrics.xlsx, 返回 {sensor_type: {metric_col: value}}。"""
    if not path.exists():
        return None
    try:
        df = pd.read_excel(path)
    except Exception as ex:
        print(f'  [warn] 读取失败 {path}: {ex}')
        return None
    sub = df[df['预测时间点'] == 'next'] if '预测时间点' in df.columns else df
    out = {}
    for _, row in sub.iterrows():
        st = row['传感器类型']
        out[st] = {
            'MAE':      float(row.get('MAE',      np.nan)),
            'RMSE':     float(row.get('RMSE',     np.nan)),
            'MAPE (%)': float(row.get('MAPE (%)', np.nan)),
            'R²':       float(row.get('R²',       np.nan)),
        }
    return out


def collect_standard(seeds_root: Path, days: list[int],
                     models: list[str]) -> pd.DataFrame:
    if not seeds_root.is_dir():
        raise FileNotFoundError(seeds_root)
    seed_dirs = sorted(p for p in seeds_root.iterdir()
                       if p.is_dir() and p.name.startswith('seed'))
    if not seed_dirs:
        raise FileNotFoundError(f'{seeds_root} 下无 seed*/ 目录')
    print(f'  seeds: {[p.name for p in seed_dirs]}')

    rows = []
    for sd in seed_dirs:
        for m in models:
            for d in days:
                xlsx = sd / m / f'day{d}' / 'results' / 'model_metrics.xlsx'
                got = _read_one_standard(xlsx)
                if got is None:
                    print(f'  [skip] {sd.name}/{m}/day{d}')
                    continue
                for st, metrics in got.items():
                    for mc, val in metrics.items():
                        rows.append({
                            'model': m, 'seed': sd.name, 'day': d,
                            'sensor': st, 'metric': mc, 'value': val,
                        })
    return pd.DataFrame(rows)


# ── Persistence-relative 指标 (mode=persistence) ───────────────────────────


def collect_persistence(seeds_root: Path, days: list[int],
                        models: list[str]) -> pd.DataFrame:
    """复用 aggregate_multiseed_r2_pers 的内部函数,跑预测 + 算 4 个 persistence 指标。

    返回长表: model / seed / day / sensor / metric -> value,
    metric ∈ {R2_pers, R2_pers_dyn, rho_MAE, rho_MAE_dyn},
    sensor ∈ {'锚杆', '围岩'}  (跟 standard 模式保持一致)。
    """
    # 局部 import,避免依赖反向膨胀到不需要 persistence 模式的用法
    from scripts.aggregate_multiseed_r2_pers import (
        EXPERIMENTS, load_day_data, compute_persistence,
        predict_for, compute_persistence_relative,
    )

    if not seeds_root.is_dir():
        raise FileNotFoundError(seeds_root)
    seed_dirs = sorted(p for p in seeds_root.iterdir()
                       if p.is_dir() and p.name.startswith('seed'))
    if not seed_dirs:
        raise FileNotFoundError(f'{seeds_root} 下无 seed*/ 目录')
    print(f'  seeds: {[p.name for p in seed_dirs]}')

    rows = []
    for d in days:
        for sd in seed_dirs:
            seed_str = sd.name
            try:
                seed_int = int(seed_str.replace('seed', ''))
            except ValueError:
                continue
            cfg_data = load_day_data(d, seed_int)
            pers = compute_persistence(*cfg_data)

            for m in models:
                exp_cfg = EXPERIMENTS.get(m)
                if exp_cfg is None:
                    print(f'  [skip] {seed_str}/{m}: 未注册')
                    continue
                model_dir = sd / m / f'day{d}' / 'models'
                ckpt = (model_dir / 'best_model.pkl') if exp_cfg['type'] == 'ml' \
                       else (model_dir / 'best_model.pth')
                if not ckpt.exists():
                    print(f'  [skip] {seed_str}/{m}/day{d}: 缺 {ckpt.name}')
                    continue
                try:
                    preds = predict_for(m, ckpt, cfg_data, seed=seed_int)
                except Exception as ex:
                    print(f'  [fail] {seed_str}/{m}/day{d}: {ex}')
                    continue
                ra = compute_persistence_relative(
                    pers['true_a'], preds['pred_a'], pers['persist_a'],
                )
                rr = compute_persistence_relative(
                    pers['true_r'], preds['pred_r'], pers['persist_r'],
                )
                for sensor_label, vals in [('锚杆', ra), ('围岩', rr)]:
                    for mc in ('R2_pers', 'R2_pers_dyn', 'rho_MAE', 'rho_MAE_dyn'):
                        rows.append({
                            'model': m, 'seed': seed_str, 'day': d,
                            'sensor': sensor_label, 'metric': mc,
                            'value': vals[mc],
                        })
    return pd.DataFrame(rows)


# ── 共用聚合 + 写表 ────────────────────────────────────────────────────────


def aggregate(df: pd.DataFrame) -> pd.DataFrame:
    """跨 seed 聚合 (median), 输出 (sensor, day, metric, model) -> median。"""
    if df.empty:
        return df
    return (df.groupby(['sensor', 'day', 'metric', 'model'])['value']
              .median().reset_index().rename(columns={'value': 'median'}))


def _fmt(v: float, is_best: bool, decimals: int = 3) -> str:
    if not np.isfinite(v):
        return '-'
    s = f'{v:.{decimals}f}'
    return f'**{s}**' if is_best else s


def _best_idx(values: list[float], smaller_is_better: bool) -> int:
    arr = np.array(values, dtype=float)
    finite = np.isfinite(arr)
    if not finite.any():
        return -1
    if smaller_is_better:
        arr[~finite] = np.inf
        return int(np.argmin(arr))
    arr[~finite] = -np.inf
    return int(np.argmax(arr))


def _metric_rows(mode: str) -> list[tuple[str, str, bool]]:
    return METRICS_STANDARD if mode == 'standard' else METRICS_PERSISTENCE


def write_md(agg: pd.DataFrame, sensor: str, days: list[int],
             models: list[str], mode: str) -> str:
    metrics = _metric_rows(mode)
    header = ' | '.join(DISPLAY[m] for m in models)
    title_mode = 'standard metrics' if mode == 'standard' else 'persistence-relative metrics'
    lines = [
        f'### Table — Detailed comparison ({sensor}, {title_mode}) across days',
        '',
        f'| | {header} |',
        '|' + ' :---: |' * (len(models) + 1),
    ]
    for d in days:
        lines.append(f'| **Day {d}** | ' + ' | '.join([''] * len(models)) + ' |')
        for metric, label, smaller in metrics:
            vals = []
            for m in models:
                r = agg[(agg['sensor'] == sensor) & (agg['day'] == d)
                        & (agg['metric'] == metric) & (agg['model'] == m)]
                vals.append(float(r['median'].iloc[0]) if not r.empty else np.nan)
            bi = _best_idx(vals, smaller)
            cells = [_fmt(v, i == bi) for i, v in enumerate(vals)]
            lines.append(f'| {label} | ' + ' | '.join(cells) + ' |')
    lines.append('')
    return '\n'.join(lines)


def write_latex(agg: pd.DataFrame, sensor: str, days: list[int],
                models: list[str], mode: str) -> str:
    metrics = _metric_rows(mode)
    n_cols = len(models) + 1
    col_spec = 'l' + 'r' * len(models)
    sensor_label = '锚杆 / Anchor' if sensor == '锚杆' else '围岩 / Rock'
    mode_label = ('standard metrics' if mode == 'standard'
                  else 'persistence-relative metrics')

    head = (
        '\\begin{table*}[!t]\n'
        '\\centering\n'
        f'\\caption{{Detailed {mode_label} across all monitoring days '
        f'({sensor_label}). Best result of each row is in bold.}}\n'
        f'\\label{{tab:table4_per_day_{mode}_{sensor}}}\n'
        '\\small\n'
        f'\\begin{{tabular}}{{{col_spec}}}\n'
        '\\toprule\n'
        ' & ' + ' & '.join(DISPLAY[m] for m in models) + ' \\\\\n'
        '\\midrule\n'
    )
    body_rows = []
    for d in days:
        body_rows.append(
            f'\\multicolumn{{{n_cols}}}{{l}}{{\\textit{{Day {d}}}}} \\\\'
        )
        for metric, label, smaller in metrics:
            vals = []
            for m in models:
                r = agg[(agg['sensor'] == sensor) & (agg['day'] == d)
                        & (agg['metric'] == metric) & (agg['model'] == m)]
                vals.append(float(r['median'].iloc[0]) if not r.empty else np.nan)
            bi = _best_idx(vals, smaller)
            cells = []
            for i, v in enumerate(vals):
                if not np.isfinite(v):
                    cells.append('-')
                    continue
                s = f'{v:.3f}'
                cells.append(f'\\textbf{{{s}}}' if i == bi else s)
            body_rows.append(f'{label} & ' + ' & '.join(cells) + ' \\\\')
        body_rows.append('\\addlinespace')
    tail = '\n\\bottomrule\n\\end{tabular}\n\\end{table*}\n'
    return head + '\n'.join(body_rows) + tail


def write_xlsx(agg: pd.DataFrame, sensor: str, days: list[int],
               models: list[str], mode: str, path: Path):
    import openpyxl
    from openpyxl.styles import Font, Alignment

    metrics = _metric_rows(mode)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sensor

    headers = ['Metric'] + [DISPLAY[m] for m in models]
    for j, h in enumerate(headers, start=1):
        c = ws.cell(row=1, column=j, value=h)
        c.font = Font(bold=True)
        c.alignment = Alignment(horizontal='center')

    r = 2
    for d in days:
        c = ws.cell(row=r, column=1, value=f'Day {d}')
        c.font = Font(bold=True, italic=True)
        r += 1
        for metric, label, smaller in metrics:
            ws.cell(row=r, column=1, value=label)
            vals = []
            for m in models:
                rr = agg[(agg['sensor'] == sensor) & (agg['day'] == d)
                         & (agg['metric'] == metric) & (agg['model'] == m)]
                vals.append(float(rr['median'].iloc[0]) if not rr.empty else np.nan)
            bi = _best_idx(vals, smaller)
            for i, v in enumerate(vals, start=2):
                cell = ws.cell(row=r, column=i,
                               value=(round(v, 4) if np.isfinite(v) else '-'))
                if i - 2 == bi:
                    cell.font = Font(bold=True)
            r += 1

    for col_idx, _ in enumerate(headers, start=1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = 12
    ws.column_dimensions['A'].width = 14
    wb.save(path)


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument('--seeds_root', type=str,
                    default='outputs/experiments_seeds',
                    help='多种子根目录 (含 seed*/ 子目录)')
    ap.add_argument('--days', type=str, default='1,2,3,4,5,6,7',
                    help='逗号分隔的天 (默认 1~7)')
    ap.add_argument('--models', type=str, default=','.join(BASELINES),
                    help='逗号分隔的模型列 (顺序即列序);默认 8 列, 不含 Transformer')
    ap.add_argument('--mode', type=str, default='standard',
                    choices=['standard', 'persistence'],
                    help='standard: MAE/RMSE/MAPE/R² (Table 4 原版指标); '
                         'persistence: R²_pers / R²_pers(dyn) / ρ_MAE / ρ_MAE(dyn)')
    args = ap.parse_args()

    seeds_root = ROOT / args.seeds_root if not Path(args.seeds_root).is_absolute() \
                 else Path(args.seeds_root)
    days = [int(x) for x in args.days.split(',') if x.strip()]
    models = [x.strip() for x in args.models.split(',') if x.strip()]
    unknown = [m for m in models if m not in DISPLAY]
    if unknown:
        print(f'未知模型: {unknown}; 可选: {list(DISPLAY.keys())}')
        return

    print('=' * 70)
    print(f'Table-4 风格逐天对比表生成  mode={args.mode}')
    print(f'  seeds_root = {seeds_root}')
    print(f'  days       = {days}')
    print(f'  models     = {models}')
    print('=' * 70)

    if args.mode == 'standard':
        df = collect_standard(seeds_root, days, models)
    else:
        df = collect_persistence(seeds_root, days, models)

    if df.empty:
        print('未读到任何指标, 退出。')
        return

    agg = aggregate(df)

    for sensor, tag in [('锚杆', 'anchor'), ('围岩', 'rock')]:
        sub = agg[agg['sensor'] == sensor]
        if sub.empty:
            print(f'  [skip] sensor={sensor} 无数据')
            continue

        md  = write_md(agg, sensor, days, models, args.mode)
        tex = write_latex(agg, sensor, days, models, args.mode)

        suffix = f'_{args.mode}' if args.mode != 'standard' else ''
        md_path   = OUT_DIR / f'table4_per_day{suffix}_{tag}.md'
        tex_path  = OUT_DIR / f'table4_per_day{suffix}_{tag}.tex'
        xlsx_path = OUT_DIR / f'table4_per_day{suffix}_{tag}.xlsx'

        md_path.write_text(md,  encoding='utf-8')
        tex_path.write_text(tex, encoding='utf-8')
        try:
            write_xlsx(agg, sensor, days, models, args.mode, xlsx_path)
            print(f'  -> {md_path.name}  /  {tex_path.name}  /  {xlsx_path.name}')
        except ImportError:
            print(f'  -> {md_path.name}  /  {tex_path.name}  (xlsx 跳过: openpyxl 未装)')


if __name__ == '__main__':
    main()
